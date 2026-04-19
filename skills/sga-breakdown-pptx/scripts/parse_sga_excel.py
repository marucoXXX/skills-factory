"""
parse_sga_excel.py — 販管費Excelファイルを読み取り、fill_sga_breakdown.py用のJSONを生成

Excelフォーマット:
  - 期ごとにシートが分かれる（シート名が期ラベルになる）
  - 各シートは「科目名」「金額」列を持つ（ヘッダー行を自動判定）
  - 売上高は別途JSON引数で指定

Usage:
  python parse_sga_excel.py \
    --excel /path/to/sga_data.xlsx \
    --sales '{"21/6期": 370, "22/6期": 424, "23/6期": 661}' \
    --output /home/claude/sga_breakdown_data.json
"""

import argparse, json, os, sys
import openpyxl

DEFAULT_COLORS = [
    "#2E7D32", "#81C784", "#4E79A7", "#1B3A5C", "#F5C242",
    "#E8C83F", "#E8923F", "#8FBC8F", "#A0522D", "#D4A0A0",
    "#B0B0B0", "#808080", "#5B9BD5", "#C0C0C0", "#70AD47",
    "#9DC3E6", "#BDD7EE", "#F4B183", "#D9D9D9", "#FFD966",
    "#4472C4", "#7B7B7B", "#ED7D31", "#A5A5A5", "#FFC000",
    "#5B9BD5", "#44546A", "#E7E6E6", "#BF8F00", "#538135"
]

NAME_KEYWORDS = ["科目", "費目", "項目", "勘定", "名称", "内訳", "account", "name"]
AMOUNT_KEYWORDS = ["金額", "額", "合計", "実績", "amount", "value", "千円", "百万円", "万円"]


def find_header_row(ws, max_scan=20):
    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        cells = [str(ws.cell(row=row_idx, column=c).value or "").strip()
                 for c in range(1, ws.max_column + 1)]
        name_col = None
        amount_col = None
        for ci, cell_text in enumerate(cells):
            low = cell_text.lower()
            if any(kw in low for kw in NAME_KEYWORDS):
                name_col = ci + 1
            if any(kw in low for kw in AMOUNT_KEYWORDS):
                amount_col = ci + 1
        if name_col and amount_col:
            return row_idx, name_col, amount_col
    return 1, 1, 2


def parse_sheet(ws):
    header_row, name_col, amount_col = find_header_row(ws)
    items = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=name_col).value
        amount = ws.cell(row=row_idx, column=amount_col).value
        if name is None or str(name).strip() == "":
            continue
        name = str(name).strip()
        if any(kw in name for kw in ["合計", "計", "Total", "total", "販管費計"]):
            continue
        if amount is None:
            amount = 0
        elif isinstance(amount, str):
            amount = amount.replace(",", "").replace("\u25b3", "-").strip()
            try:
                amount = float(amount)
            except ValueError:
                amount = 0
        else:
            amount = float(amount)
        items.append({"name": name, "amount": round(amount)})
    return items


def build_json(excel_path, sales_data, main_message="", chart_title="",
               source="", unit_label="", threshold=5.0):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_data = {}
    all_categories = []
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        items = parse_sheet(ws)
        sheet_data[ws_name] = items
        for item in items:
            if item["name"] not in all_categories:
                all_categories.append(item["name"])

    categories = []
    for i, cat_name in enumerate(all_categories):
        color = DEFAULT_COLORS[i % len(DEFAULT_COLORS)]
        categories.append({"name": cat_name, "color": color})

    periods = []
    for ws_name in wb.sheetnames:
        items = sheet_data[ws_name]
        item_map = {it["name"]: it["amount"] for it in items}
        values = [item_map.get(cat_name, 0) for cat_name in all_categories]
        total = sum(v for v in values if v > 0)
        sales = sales_data.get(ws_name, 0)
        line_value = round(total / sales * 100, 1) if sales > 0 else 0
        periods.append({
            "label": ws_name,
            "values": values,
            "total": total,
            "line_value": line_value
        })

    result = {
        "main_message": main_message or "販管費の構成比推移と対売上高比率の変化を把握すべき",
        "chart_title": chart_title or "対売上販管費比率と、販管費内構成比の推移",
        "source": source or "対象会社提供資料",
        "unit_label": unit_label or "（単位：百万円、%）",
        "label_threshold_pct": threshold,
        "categories": categories,
        "line": {"series_name": "対売上高販管費比率", "color": "#666666"},
        "periods": periods,
        "trend_arrow": True
    }
    wb.close()
    return result


def main():
    ap = argparse.ArgumentParser(description="Parse SGA Excel to JSON")
    ap.add_argument("--excel", required=True)
    ap.add_argument("--sales", required=True,
                    help='Sales per period as JSON string')
    ap.add_argument("--output", required=True)
    ap.add_argument("--main-message", default="")
    ap.add_argument("--chart-title", default="")
    ap.add_argument("--source", default="")
    ap.add_argument("--unit-label", default="")
    ap.add_argument("--threshold", type=float, default=5.0)
    args = ap.parse_args()

    sales_data = json.loads(args.sales)
    result = build_json(args.excel, sales_data,
                        main_message=args.main_message,
                        chart_title=args.chart_title,
                        source=args.source,
                        unit_label=args.unit_label,
                        threshold=args.threshold)

    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"JSON saved: {args.output}")
    print(f"  Periods: {len(result['periods'])}")
    print(f"  Categories: {len(result['categories'])}")
    for p in result['periods']:
        print(f"    {p['label']}: total={p['total']}, ratio={p['line_value']}%")


if __name__ == "__main__":
    main()
